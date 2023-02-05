import csv
import difflib
import os
import traceback
import typing

import openpyxl
from openpyxl.workbook import Workbook

from my_types import RowDataDict


def sanitize_cell_value(value: str) -> str:
    """
    Sanitizes a cell value by removing all spaces and newlines
    :param value: The value to sanitize
    :return: The sanitized value
    """
    return clean_value(sanitize_names(value))


def lower_clean_cell_value(value: str) -> str:
    """
    Sanitizes a cell value by removing all spaces and newlines and lower casing the value

    :param value: The value to sanitize
    :return:  The sanitized value
    """
    return sanitize_cell_value(value).lower()


def clean_value(value: str) -> str:
    """
    Remove all spaces and newlines from the value.
    :param value: The value to clean
    :return: The cleaned value
    """
    if value is None:
        return ''
    return value.strip().replace("\n", "") if isinstance(value, str) else value


def sanitize_names(name: typing.Any) -> str:
    """
    Sanitizes a name by removing all spaces around and special characters.
    :param name: The name to sanitize will be cast to a string
    :return:  The sanitized name
    """
    sanitized_name = str(name)

    if not sanitized_name:
        return ''
    if sanitized_name[0] == '\xa0':
        sanitized_name = sanitized_name.lstrip("\xa0")
    if sanitized_name[-1] == '\xa0':
        sanitized_name = sanitized_name.rstrip('\xa0')
    if "\xa0" in sanitized_name:
        sanitized_name = sanitized_name.replace("\xa0", " ")
    if sanitized_name[0] == " ":
        sanitized_name = sanitized_name.lstrip(" ")
    return sanitized_name


def transform_csv_to_xslx(file_path: str) -> Workbook:
    """
    Takes a file and creates an openpyxl Workbook object from it

    Accepts csv, tsv, and txt files

    :param file_path: The path to the file
    :return: A Workbook object
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter='\t')
        wb = openpyxl.Workbook()
        ws = wb.active
        # Have to use a while loop because the file has a lot of UnicodeDecodeErrors,
        # and we cannot capture errors during a for loop
        while True:
            try:
                try:
                    row = next(reader)
                except UnicodeDecodeError:
                    continue
                ws.append(row)
            except StopIteration:
                break
        return wb


def read_xslx_file(xlsx_file: str | Workbook, **kwargs) -> typing.List[dict]:
    """
    Takes a xlsx file and returns a Workbook object
    :param xlsx_file: The path to the file or a Workbook object
    :param kwargs:  Any additional arguments to pass to the load_workbook function
    :return:  A list of dictionaries representing the rows in the file
    """
    if isinstance(xlsx_file, str):
        wb = openpyxl.load_workbook(xlsx_file, **kwargs)
    else:
        wb = xlsx_file
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet(title='Worksheet1', index=0) if len(wb.worksheets) == 0 else wb.worksheets[0]
    rows = list(ws.rows)
    header = [sanitize_names(cell.value) for cell in rows[0]]
    try:
        result = [
            dict(zip(header, [clean_value(cell.value) for cell in row]))
            for row in rows[1:]
        ]
        print("✅")
        return result
    except Exception as error:
        print("❌")
        traceback.print_tb(error.__traceback__)


def find_file(file_name: str) -> typing.Optional[str]:
    """
    Finds the closest matching file in the current directory
    :param file_name: The name of the file to find
    :return: The path to the file
    """
    for root, dirs, files in os.walk("./files"):
        if closest_match := difflib.get_close_matches(file_name, files, n=1):
            closest_match: typing.List[str]
            return os.path.join(root, closest_match[0])
    return None


def read_report(file_name: str, **kwargs) -> typing.List[dict]:
    """
    Reads a report file and returns a list of dictionaries
    :param file_name: The name of the file to read
    :param kwargs: Any additional arguments to pass to the read_xslx_file function
    :return: A list of dictionaries representing the rows in the file
    """
    file_path = find_file(file_name)
    print(f"Reading {file_path}", end=' ')
    if not file_path:
        raise FileNotFoundError(f"Could not find file {file_name}")

    if file_path.endswith((".xlsx", ".xls")):
        return read_xslx_file(file_path, **kwargs)
    elif file_path.endswith((".csv", ".txt", '.tsv')):
        wb = transform_csv_to_xslx(file_path)
        return read_xslx_file(wb, **kwargs)


def generate_mapped_cell_dict() -> dict:
    """
    Generates a template dictionary from OUTPUT_MAPPED_CELLS column_name
    :return: A dictionary with the column names as keys and empty strings as values
    """
    from cell_mapping import OUTPUT_MAPPED_CELLS
    return {cell['column_name']: '' for cell in OUTPUT_MAPPED_CELLS}


def generate_row_data_dict() -> RowDataDict:
    """
    Generates a template dictionary for row data
    :return:  A dictionary with the keys inventory_row, informed_row, and restock_row
    """
    return {'inventory_row': {}, 'informed_row': {}, 'restock_row': {}}


def pick_marketplace():
    """Picks marketplace from user input"""
    while True:
        marketplace = input("Enter marketplace ID or (quit to exit): ")
        if marketplace == 'quit':
            exit(0)
        if not marketplace:
            print("Please enter a marketplace ID")
            continue
        if not marketplace.isdigit():
            print("Please enter a valid marketplace ID")
            continue
        return int(marketplace)


